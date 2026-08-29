"""
Drosophila Movement Tracker - GUI Application
______________________________________________

Read README.md for details

Environemnt: PyQt5 

Export filenames are generated automatically and are not user-editable:
  - Excel workbook : "<session name>_<date>_<time>.xlsx"
  - Trajectory PNGs/Graphs: "<session name>_<date>_<time>_<fly name>.png"

Dependencies:
    pip install PyQt5 opencv-python openpyxl matplotlib numpy
    pip install pygrabber   # Windows only, optional - enables real camera
                             # device names instead of generic "Camera N"

Run:
    python fly_tracker_gui.py
"""

import sys
import os
import json
import time
import datetime
from dataclasses import dataclass, field

os.environ["OPENCV_LOG_LEVEL"] = "SILENT"

import cv2
import numpy as np
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QObject
from PyQt5.QtGui import QImage, QPixmap, QPainter, QPen, QColor
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QPushButton,
    QComboBox, QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox, QSlider,
    QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox, QCheckBox, QColorDialog, 
    QSpinBox, QDoubleSpinBox, QLineEdit, QScrollArea, QProgressBar,
    QRadioButton, QButtonGroup
)

from openpyxl import Workbook
from openpyxl.styles import PatternFill
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt


DEFAULT_NUM_FLIES = 5
MAX_FLIES = 100
MAX_CAMERA_PROBE = 8  # how many camera indices to test when scanning
FLY_COLORS = [
    "#e6194B", "#3cb44b", "#4363d8", "#f58231", "#911eb4",
    "#42d4f4", "#f032e6", "#bfef45", "#fabed4", "#469990",
    "#dcbeff", "#9A6324", "#fffac8", "#800000", "#aaffc3",
    "#808000", "#ffd8b1", "#000075", "#a9a9a9", "#800080",
]
DEFAULT_LOG_INTERVAL_SEC = 1.0
DEFAULT_VIAL_W_MM = 105.0
DEFAULT_VIAL_H_MM = 105.0
# (label, width, height) - 16:9 modes listed first since that's usually
# what's wanted; 4:3 kept available since some cameras only support that.
RESOLUTIONS = [
    ("1920x1080 (16:9)", 1920, 1080),
    ("1280x720 (16:9)", 1280, 720),
    ("960x540 (16:9)", 960, 540),
    ("640x480 (4:3)", 640, 480),
    ("800x600 (4:3)", 800, 600),
]
DEFAULT_RESOLUTION_INDEX = 0  # 1920 X 1080
FPS_OPTIONS = [15, 24, 30, 60]
DEFAULT_FPS_INDEX = 3  # 60 -- OBS Virtual Camera or similar sources can go higher (select 60)
DEFAULT_MOVEMENT_THRESHOLD_MM = 0.5
DEFAULT_GRID_DIVISIONS = 4
IDLE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red
MOVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green


def fly_color(idx):
    return FLY_COLORS[idx % len(FLY_COLORS)]

def get_fly_color(roi, idx):
    return roi.color if getattr(roi, "color", None) else fly_color(idx)

def safe_filename_part(name):
    return "".join(c if c.isalnum() or c in " _-" else "_" for c in name).strip() or "fly"


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class FlyROI:
    name: str = "Fly"
    rect: tuple = None  # (x, y, w, h) in raw frame pixel coords, or None
    real_w_mm: float = DEFAULT_VIAL_W_MM
    real_h_mm: float = DEFAULT_VIAL_H_MM
    color: str = None


@dataclass
class FlyTrack:
    timestamps: list = field(default_factory=list)
    xs_mm: list = field(default_factory=list)
    ys_mm: list = field(default_factory=list)
    distances_mm: list = field(default_factory=list)  # cumulative total distance at each logged row
    last_seen: float = None
    last_pos_mm: tuple = None
    smoothing_buffer: list = field(default_factory=list)
    total_distance_mm: float = 0.0
    last_logged_pos: tuple = None  # last logged (x, y) used to accumulate distance


# ---------------------------------------------------------------------------
# Tracker engine: camera/video I/O, detection, logging, export
# ---------------------------------------------------------------------------

class TrackerEngine(QObject):
    frame_ready = pyqtSignal(np.ndarray)
    status_updated = pyqtSignal(int, object, object)  # fly_idx, (x,y)mm or None, detected
    distance_updated = pyqtSignal(int, float)  # fly_idx, cumulative total distance (mm)
    fps_updated = pyqtSignal(float)
    progress_updated = pyqtSignal(int)  # 0-100, used during video-file batch processing

    def __init__(self, num_flies=DEFAULT_NUM_FLIES):
        super().__init__()
        self.cap = None
        self.threshold = 90
        self.min_blob_area = 8
        self.log_interval_sec = DEFAULT_LOG_INTERVAL_SEC
        self.smoothing_window = 1  # 1 = no smoothing
        self.preview_mask = False
        self.movement_threshold_mm = DEFAULT_MOVEMENT_THRESHOLD_MM
        self.grid_divisions = DEFAULT_GRID_DIVISIONS
        self.rois = []
        self.tracks = []
        self.set_num_flies(num_flies)
        self.tracking_active = False
        self.last_log_time = 0.0
        self._last_frame_time = None
        self._cancel_batch = False

    # -- camera discovery -------------------------------------------------------
    @staticmethod
    def detect_cameras(max_probe=MAX_CAMERA_PROBE):
        """Return a list of (index, name) tuples for cameras actually connected to the system.
        """
        device_names = {}
        try:
            from pygrabber.dshow_graph import FilterGraph
            graph = FilterGraph()
            for i, name in enumerate(graph.get_input_devices()):
                device_names[i] = name
        except Exception:
            pass  # pygrabber not available - fall back to generic names below

        found = []
        for i in range(max_probe):
            backend = cv2.CAP_DSHOW if sys.platform.startswith("win") else cv2.CAP_V4L2
            cap = cv2.VideoCapture(i, backend)
            if cap is not None and cap.isOpened():
                found.append((i, device_names.get(i, f"Camera {i}")))
            cap.release()
        return found

    # -- fly count management -------------------------------------------------
    def set_num_flies(self, n):
        """Resize the ROI/track lists to n flies, preserving existing entries (ROI position, name, size, logged data) for indices that still exist."""
        n = max(1, min(MAX_FLIES, n))
        old_rois = self.rois
        old_tracks = self.tracks
        new_rois = []
        new_tracks = []
        for i in range(n):
            if i < len(old_rois):
                new_rois.append(old_rois[i])
                new_tracks.append(old_tracks[i])
            else:
                new_rois.append(FlyROI(name=f"Fly {i+1}"))
                new_tracks.append(FlyTrack())
        self.rois = new_rois
        self.tracks = new_tracks

    @property
    def num_flies(self):
        return len(self.rois)

    def set_fly_name(self, idx, name):
        self.rois[idx].name = name.strip() or f"Fly {idx+1}"

    # -- source lifecycle -----------------------------------------------------
    def open_video_source(self, index=0, width=1280, height=720, fps=30):
        """Open a live camera, explicitly requesting a resolution, frame rate, and pixel format.
        Without setting these, cv2.VideoCapture default is 640x480 (4:3) at a throttled frame rate - even on cameras that support 720p/30fps.
        """
        backend = cv2.CAP_DSHOW if sys.platform.startswith("win") else cv2.CAP_ANY
        self.cap = cv2.VideoCapture(index, backend)
        if not self.cap.isOpened():
            raise RuntimeError("Could not open camera. Check the camera index/connection.")
        self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*"MJPG"))
        self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, width)
        self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, height)
        self.cap.set(cv2.CAP_PROP_FPS, fps)
        # Some cameras/drivers clamp to the nearest mode they actually support
        # rather than the exact values requested - read back what we got.
        actual_w = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        actual_h = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        actual_fps = self.cap.get(cv2.CAP_PROP_FPS)
        return actual_w, actual_h, actual_fps

    def open_video_file(self, path):
        cap = cv2.VideoCapture(path)
        if not cap.isOpened():
            raise RuntimeError(f"Could not open video file: {path}")
        return cap

    def close_video_source(self):
        if self.cap is not None:
            self.cap.release()
            self.cap = None

    # -- setters ---------------------------------------------------------------
    def set_threshold(self, value):
        self.threshold = value

    def set_min_blob_area(self, value):
        self.min_blob_area = value

    def set_log_interval(self, seconds):
        self.log_interval_sec = seconds

    def set_smoothing_window(self, n):
        self.smoothing_window = max(1, n)

    def set_preview_mask(self, enabled):
        self.preview_mask = enabled

    def set_movement_threshold(self, mm):
        self.movement_threshold_mm = mm

    def set_grid_divisions(self, n):
        self.grid_divisions = max(0, n)

    def set_roi(self, fly_idx, rect):
        self.rois[fly_idx].rect = rect

    def clear_roi(self, fly_idx):
        self.rois[fly_idx].rect = None

    def set_vial_size(self, fly_idx, w_mm, h_mm):
        self.rois[fly_idx].real_w_mm = w_mm
        self.rois[fly_idx].real_h_mm = h_mm 

    def set_fly_color(self, fly_idx, hex_color):
        self.rois[fly_idx].color = hex_color

    def reset_data(self):
        self.tracks = [FlyTrack() for _ in range(self.num_flies)]
        self.last_log_time = 0.0

    # -- layout save/load -------------------------------------------------------
    def save_layout(self, path):
        data = {
            "rois": [
                {"name": r.name, "rect": r.rect, "real_w_mm": r.real_w_mm, "real_h_mm": r.real_h_mm, "color": r.color}
                for r in self.rois
            ],
            "threshold": self.threshold,
            "min_blob_area": self.min_blob_area,
            "log_interval_sec": self.log_interval_sec,
            "smoothing_window": self.smoothing_window,
        }
        with open(path, "w") as f:
            json.dump(data, f, indent=2)

    def load_layout(self, path):
        with open(path) as f:
            data = json.load(f)
        roi_data = data.get("rois", [])
        self.set_num_flies(len(roi_data) or self.num_flies)
        for i, r in enumerate(roi_data):
            rect = r.get("rect")
            self.rois[i].name = r.get("name", f"Fly {i+1}")
            self.rois[i].rect = tuple(rect) if rect else None
            self.rois[i].real_w_mm = r.get("real_w_mm", DEFAULT_VIAL_W_MM)
            self.rois[i].real_h_mm = r.get("real_h_mm", DEFAULT_VIAL_H_MM)
            self.rois[i].color = r.get("color")
        self.threshold = data.get("threshold", self.threshold)
        self.min_blob_area = data.get("min_blob_area", self.min_blob_area)
        self.log_interval_sec = data.get("log_interval_sec", self.log_interval_sec)
        self.smoothing_window = data.get("smoothing_window", self.smoothing_window)

    # -- shared per-frame detection (used by both live + batch video modes) ------
    def _detect_on_frame(self, frame, sim_time, should_log):
        """Runs detection for every fly and ALWAYS draws the ROI box + a marker
        at the detected position on `display`, regardless of should_log/
        tracking_active - the overlay is a permanent live aid, not tied to
        whether data logging is currently switched on."""
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        display = frame.copy()

        for i, roi in enumerate(self.rois):
            detected = False
            pos_mm = None

            if roi.rect is not None:
                x, y, w, h = (int(v) for v in roi.rect)
                sub = gray[max(y, 0):y + h, max(x, 0):x + w]

                if sub.size > 0:
                    _, mask = cv2.threshold(sub, self.threshold, 255, cv2.THRESH_BINARY_INV)
                    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    if contours:
                        largest = max(contours, key=cv2.contourArea)
                        if cv2.contourArea(largest) > self.min_blob_area:
                            M = cv2.moments(largest)
                            if M["m00"] != 0:
                                cx = M["m10"] / M["m00"]
                                cy = M["m01"] / M["m00"]
                                x_mm = (cx / w) * roi.real_w_mm
                                y_mm = (cy / h) * roi.real_h_mm
                                pos_mm = self._apply_smoothing(i, x_mm, y_mm)
                                detected = True
                                # Marker on the detected fly position - a yellow X-shaped crosshair, always drawn.
                                mx, my = x + int(cx), y + int(cy)
                                arm = 4
                                cv2.line(display, (mx - arm, my - arm), (mx + arm, my + arm),
                                         (0, 255, 255), 1, lineType=cv2.LINE_AA)
                                cv2.line(display, (mx - arm, my + arm), (mx + arm, my - arm),
                                         (0, 255, 255), 1, lineType=cv2.LINE_AA)

                    if self.preview_mask:
                        mask_bgr = cv2.cvtColor(mask, cv2.COLOR_GRAY2BGR)
                        display[max(y, 0):y + h, max(x, 0):x + w] = mask_bgr

                # ROI rectangle + label - always drawn whenever an ROI is set
                color = tuple(int(c) for c in QColor(get_fly_color(roi, i)).getRgb()[2::-1])
                cv2.rectangle(display, (x, y), (x + w, y + h), color, 2)
                cv2.putText(display, roi.name, (x, max(y - 6, 10)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)

                # Secondary grid inside the ROI - visual aid for judging position/ scale at a glance, purely cosmetic (does not affect detection).
                if self.grid_divisions > 1:
                    grid_color = (110, 110, 110)
                    for g in range(1, self.grid_divisions):
                        gx = x + int(w * g / self.grid_divisions)
                        gy = y + int(h * g / self.grid_divisions)
                        cv2.line(display, (gx, y), (gx, y + h), grid_color, 1)
                        cv2.line(display, (x, gy), (x + w, gy), grid_color, 1)

            track = self.tracks[i]
            if detected:
                track.last_seen = sim_time
                track.last_pos_mm = pos_mm

            if should_log:
                track.timestamps.append(datetime.datetime.now())
                track.xs_mm.append(pos_mm[0] if detected else None)
                track.ys_mm.append(pos_mm[1] if detected else None)

                # Total distance moved: accumulated as soon as tracking starts, using the straight-line (Euclidean) distance between this logged point and the previously logged point, in mm - but ONLY counted when the X change, the Y change, or both are at/above the movement threshold 
                if detected:
                    if track.last_logged_pos is not None:
                        dx = pos_mm[0] - track.last_logged_pos[0]
                        dy = pos_mm[1] - track.last_logged_pos[1]
                        if abs(dx) > self.movement_threshold_mm or abs(dy) > self.movement_threshold_mm:
                            track.total_distance_mm += (dx ** 2 + dy ** 2) ** 0.5
                    track.last_logged_pos = pos_mm
                track.distances_mm.append(track.total_distance_mm)
                self.distance_updated.emit(i, track.total_distance_mm)

            self.status_updated.emit(i, pos_mm, detected)

        return display

    # -- live per-frame processing (driven by QTimer) -----------------------------
    def process_frame(self):
        if self.cap is None:
            return
        ok, frame = self.cap.read()
        if not ok:
            return

        now = time.time()
        if self._last_frame_time is not None:
            dt = now - self._last_frame_time
            if dt > 0:
                self.fps_updated.emit(1.0 / dt)
        self._last_frame_time = now

        should_log = self.tracking_active and (now - self.last_log_time) >= self.log_interval_sec
        display = self._detect_on_frame(frame, now, should_log)

        if should_log:
            self.last_log_time = now

        self.frame_ready.emit(display)

    def _apply_smoothing(self, fly_idx, x_mm, y_mm):
        track = self.tracks[fly_idx]
        buf = track.smoothing_buffer
        buf.append((x_mm, y_mm))
        if len(buf) > self.smoothing_window:
            buf.pop(0)
        avg_x = sum(p[0] for p in buf) / len(buf)
        avg_y = sum(p[1] for p in buf) / len(buf)
        return (avg_x, avg_y)

    # -- batch processing of a video file ------------------------------------------
    def cancel_batch(self):
        self._cancel_batch = True

    def process_video_batch(self, path):
        """Runs synchronously, emitting progress/frame/status signals as it goes.
        ."""
        self._cancel_batch = False
        cap = self.open_video_file(path)
        fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) or 1
        last_logged_video_time = -self.log_interval_sec

        frame_idx = 0
        while True:
            if self._cancel_batch:
                break
            ok, frame = cap.read()
            if not ok:
                break

            video_time = frame_idx / fps
            should_log = (video_time - last_logged_video_time) >= self.log_interval_sec
            display = self._detect_on_frame(frame, video_time, should_log)
            if should_log:
                last_logged_video_time = video_time

            self.frame_ready.emit(display)
            self.progress_updated.emit(int(100 * frame_idx / total_frames))

            frame_idx += 1
            yield  # let the caller pump the Qt event loop between frames

        cap.release()
        self.progress_updated.emit(100)

    # -- export ---------------------------------------------------------------
    def export_excel(self, path):
        wb = Workbook()
        wb.remove(wb.active)
        used_names = set()
        for i, (roi, track) in enumerate(zip(self.rois, self.tracks)):
            sheet_name = roi.name[:31] or f"Fly {i+1}"  # Excel sheet name limit
            base_name, suffix = sheet_name, 1
            while sheet_name in used_names:
                suffix += 1
                sheet_name = f"{base_name}_{suffix}"[:31]
            used_names.add(sheet_name)

            ws = wb.create_sheet(title=sheet_name)
            ws.append([
                "Timestamp", "X (mm)", "Y (mm)", "Status (0=idle, 1=move)",
                "Instantaneous Speed (mm/s)", "Avg Speed - Move Segment (mm/s)",
                "Total Distance (mm)",
            ])

            n = len(track.timestamps)
            statuses = [None] * n
            dist_steps = [None] * n   # raw per-step distance (not threshold-gated)
            dt_steps = [None] * n     # seconds since the previous logged point
            inst_speeds = [None] * n  # dist_steps / dt_steps

            prev_x, prev_y, prev_t = None, None, None
            for idx in range(n):
                t, x, y = track.timestamps[idx], track.xs_mm[idx], track.ys_mm[idx]
                if x is not None and y is not None and prev_x is not None and prev_y is not None:
                    dx, dy = x - prev_x, y - prev_y
                    dist = (dx ** 2 + dy ** 2) ** 0.5
                    dt = (t - prev_t).total_seconds() if prev_t is not None else None
                    # 1 (move) if either X or Y change is strictly greater than
                    # the threshold; 0 (idle) otherwise - same rule used for
                    # total-distance accumulation, kept consistent here.
                    statuses[idx] = 1 if (abs(dx) > self.movement_threshold_mm or abs(dy) > self.movement_threshold_mm) else 0
                    dist_steps[idx] = dist
                    dt_steps[idx] = dt
                    if dt and dt > 0:
                        inst_speeds[idx] = dist / dt
                if x is not None and y is not None:
                    prev_x, prev_y, prev_t = x, y, t

            # Average speed per continuous "move" segment: a maximal run of
            # consecutive status==1 rows. Duration/distance are summed across
            # that run's steps (each step already spans from the previous
            # logged point, so this naturally covers "from when it started
            # moving to when it went idle"); the resulting average is written
            # on every row of that segment.
            seg_avg = [None] * n
            idx = 0
            while idx < n:
                if statuses[idx] == 1:
                    start = idx
                    while idx < n and statuses[idx] == 1:
                        idx += 1
                    end = idx  # exclusive
                    seg_dist = sum(d for d in dist_steps[start:end] if d is not None)
                    seg_time = sum(d for d in dt_steps[start:end] if d is not None)
                    avg = seg_dist / seg_time if seg_time and seg_time > 0 else None
                    for k in range(start, end):
                        seg_avg[k] = avg
                else:
                    idx += 1

            for row_idx in range(n):
                excel_row = row_idx + 2
                t, x, y = track.timestamps[row_idx], track.xs_mm[row_idx], track.ys_mm[row_idx]
                status = statuses[row_idx]
                inst_speed = inst_speeds[row_idx]
                avg_speed = seg_avg[row_idx]
                cum_dist = track.distances_mm[row_idx] if row_idx < len(track.distances_mm) else None

                ws.append([
                    t.strftime("%Y-%m-%d %H:%M:%S"), x, y, status,
                    round(inst_speed, 3) if inst_speed is not None else None,
                    round(avg_speed, 3) if avg_speed is not None else None,
                    round(cum_dist, 3) if cum_dist is not None else None,
                ])
                if status == 0:
                    ws.cell(row=excel_row, column=4).fill = IDLE_FILL
                elif status == 1:
                    ws.cell(row=excel_row, column=4).fill = MOVE_FILL

        wb.save(path)

    def export_trajectories(self, out_dir, filename_prefix):
        """Saves TWO PNGs per fly: X-position-vs-time and Y-position-vs-time,
        named '<filename_prefix>_<fly name>_X_vs_time.png' and
        '..._Y_vs_time.png'."""
        for i, (roi, track) in enumerate(zip(self.rois, self.tracks)):
            valid = [
                (t, x, y) for t, x, y in zip(track.timestamps, track.xs_mm, track.ys_mm)
                if x is not None and y is not None
            ]
            if not valid:
                continue
            t0 = valid[0][0]
            times_s = [(t - t0).total_seconds() for t, x, y in valid]
            xs = [x for t, x, y in valid]
            ys = [y for t, x, y in valid]
            safe_name = safe_filename_part(roi.name)

            fig, ax = plt.subplots(figsize=(6, 3))
            ax.plot(times_s, xs, "-", color=get_fly_color(roi, i), linewidth=1)
            ax.set_xlabel("Time (s)")
            ax.set_ylabel("X (mm)")
            ax.set_title(f"{roi.name} - X position vs Time")
            fig.tight_layout()
            fig.savefig(os.path.join(out_dir, f"{filename_prefix}_{safe_name}_X_vs_time.png"), dpi=150)
            plt.close(fig)

            fig, ax = plt.subplots(figsize=(6, 3))
            ax.plot(times_s, ys, "-", color=get_fly_color(roi, i), linewidth=1)
            ax.set_xlabel("Time (s)")
            ax.set_ylabel("Y (mm)")
            ax.set_title(f"{roi.name} - Y position vs Time")
            fig.tight_layout()
            fig.savefig(os.path.join(out_dir, f"{filename_prefix}_{safe_name}_Y_vs_time.png"), dpi=150)
            plt.close(fig)


# ---------------------------------------------------------------------------
# Video label with click-and-drag ROI drawing
# ---------------------------------------------------------------------------

class VideoLabel(QLabel):
    roi_drawn = pyqtSignal(tuple)

    def __init__(self):
        super().__init__()
        self.setMinimumSize(560, 420)
        self.setStyleSheet("background-color: #202020;")
        self.setAlignment(Qt.AlignCenter)
        self._drawing = False
        self._start_pt = None
        self._end_pt = None
        self.frame_size = (640, 480)
        self.roi_mode = False

    def arm_roi_mode(self, armed: bool):
        self.roi_mode = armed

    def mousePressEvent(self, event):
        if self.roi_mode and event.button() == Qt.LeftButton:
            self._drawing = True
            self._start_pt = event.pos()
            self._end_pt = event.pos()

    def mouseMoveEvent(self, event):
        if self._drawing:
            self._end_pt = event.pos()
            self.update()

    def mouseReleaseEvent(self, event):
        if self._drawing and event.button() == Qt.LeftButton:
            self._drawing = False
            self._end_pt = event.pos()
            rect = self._label_rect_to_frame_rect(self._start_pt, self._end_pt)
            if rect is not None:
                self.roi_drawn.emit(rect)
            self.roi_mode = False
            self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        if self._drawing and self._start_pt and self._end_pt:
            painter = QPainter(self)
            pen = QPen(QColor("#ffff00"))
            pen.setWidth(2)
            painter.setPen(pen)
            x1, y1 = self._start_pt.x(), self._start_pt.y()
            x2, y2 = self._end_pt.x(), self._end_pt.y()
            painter.drawRect(min(x1, x2), min(y1, y2), abs(x2 - x1), abs(y2 - y1))

    def _label_rect_to_frame_rect(self, p1, p2):
        pixmap = self.pixmap()
        if pixmap is None:
            return None
        lw, lh = self.width(), self.height()
        pw, ph = pixmap.width(), pixmap.height()
        offset_x = (lw - pw) / 2
        offset_y = (lh - ph) / 2

        def to_frame(pt):
            fx = (pt.x() - offset_x) / pw * self.frame_size[0]
            fy = (pt.y() - offset_y) / ph * self.frame_size[1]
            return fx, fy

        fx1, fy1 = to_frame(p1)
        fx2, fy2 = to_frame(p2)
        x, y = min(fx1, fx2), min(fy1, fy2)
        w, h = abs(fx2 - fx1), abs(fy2 - fy1)
        if w < 5 or h < 5:
            return None
        return (x, y, w, h)


# ---------------------------------------------------------------------------
# Main window - single page, no tabs, everything in a scrollable side panel
# ---------------------------------------------------------------------------

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Drosophila Movement Tracker")
        self.engine = TrackerEngine(DEFAULT_NUM_FLIES)
        self.engine.frame_ready.connect(self.on_frame)
        self.engine.status_updated.connect(self.on_status)
        self.engine.distance_updated.connect(self.on_distance_updated)
        self.engine.fps_updated.connect(self.on_fps)
        self.engine.progress_updated.connect(self.on_progress)

        self.timer = QTimer()
        self.timer.timeout.connect(self.engine.process_frame)

        self.session_start_time = None
        self.duration_limit_sec = 0
        self.auto_save_enabled = False
        self.export_folder = os.getcwd()
        self.loaded_video_path = None
        self._batch_generator = None
        self._selected_fly_idx = 0
        self._suppress_table_signal = False

        self.elapsed_timer = QTimer()
        self.elapsed_timer.timeout.connect(self.update_elapsed)

        self._build_ui()
        self.rebuild_status_table()
        self.on_mode_changed()
        self.refresh_camera_list()

    # -- UI construction ----------------------------------------------------
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)

        # Left: video preview + live status + progress bar
        left_layout = QVBoxLayout()
        self.video_label = VideoLabel()
        self.video_label.roi_drawn.connect(self.on_roi_drawn)
        left_layout.addWidget(self.video_label, stretch=3)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        left_layout.addWidget(self.progress_bar)

        status_group = QGroupBox("Live Status  (double-click Fly name to rename; click a row to select it for ROI setup)")
        self.status_layout = QVBoxLayout(status_group)
        self.status_table = QTableWidget(0, 4)
        self.status_table.setHorizontalHeaderLabels(["Fly", "Detected", "X, Y (mm)", "Total Distance (mm)"])
        self.status_table.itemChanged.connect(self.on_table_item_changed)
        self.status_table.itemSelectionChanged.connect(self.on_table_selection_changed)
        self.status_layout.addWidget(self.status_table)

        info_row = QHBoxLayout()
        self.fps_label = QLabel("Frame Rate: -")
        self.elapsed_label = QLabel("Elapsed Time: 00:00:00")
        info_row.addWidget(self.fps_label)
        info_row.addWidget(self.elapsed_label)
        self.status_layout.addLayout(info_row)

        left_layout.addWidget(status_group, stretch=1)
        main_layout.addLayout(left_layout, stretch=3)

        # Right: single scrollable panel with every control, no tabs
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedWidth(360)
        panel = QWidget()
        panel_layout = QVBoxLayout(panel)

        panel_layout.addWidget(self._build_mode_group())
        panel_layout.addWidget(self._build_camera_group())
        panel_layout.addWidget(self._build_flies_group())
        panel_layout.addWidget(self._build_roi_group())
        panel_layout.addWidget(self._build_detection_group())
        panel_layout.addWidget(self._build_session_group())
        panel_layout.addWidget(self._build_export_group())
        panel_layout.addStretch()

        scroll.setWidget(panel)
        main_layout.addWidget(scroll, stretch=2)

    def _build_mode_group(self):
        group = QGroupBox("Source Mode")
        layout = QGridLayout(group)

        self.mode_group_btns = QButtonGroup(self)
        self.live_mode_radio = QRadioButton("Live Camera")
        self.video_mode_radio = QRadioButton("Video File")
        self.live_mode_radio.setChecked(True)
        self.mode_group_btns.addButton(self.live_mode_radio)
        self.mode_group_btns.addButton(self.video_mode_radio)
        self.live_mode_radio.toggled.connect(self.on_mode_changed)
        layout.addWidget(self.live_mode_radio, 0, 0)
        layout.addWidget(self.video_mode_radio, 0, 1)

        self.load_video_btn = QPushButton("Load Video")
        self.load_video_btn.clicked.connect(self.load_video_file)
        layout.addWidget(self.load_video_btn, 1, 0)

        self.process_video_btn = QPushButton("Process Video")
        self.process_video_btn.clicked.connect(self.process_video)
        layout.addWidget(self.process_video_btn, 1, 1)

        self.video_path_label = QLabel("No video loaded")
        self.video_path_label.setWordWrap(True)
        layout.addWidget(self.video_path_label, 2, 0, 1, 2)

        return group

    def _build_camera_group(self):
        group = QGroupBox("Camera")
        layout = QGridLayout(group)

        layout.addWidget(QLabel("Camera"), 0, 0)
        self.camera_select = QComboBox()
        layout.addWidget(self.camera_select, 0, 1)

        layout.addWidget(QLabel("Resolution"), 1, 0)
        self.resolution_select = QComboBox()
        self.resolution_select.addItems([r[0] for r in RESOLUTIONS])
        self.resolution_select.setCurrentIndex(DEFAULT_RESOLUTION_INDEX)
        layout.addWidget(self.resolution_select, 1, 1)

        layout.addWidget(QLabel("Frame Rate"), 2, 0)
        self.fps_select = QComboBox()
        self.fps_select.addItems([f"{f} fps" for f in FPS_OPTIONS])
        self.fps_select.setCurrentIndex(DEFAULT_FPS_INDEX)
        self.fps_select.setToolTip("Request a higher rate (e.g. 60) if using a source that supports it, such as OBS Virtual Camera.")
        layout.addWidget(self.fps_select, 2, 1)

        self.refresh_cam_btn = QPushButton("Refresh Cameras")
        self.refresh_cam_btn.clicked.connect(self.refresh_camera_list)
        layout.addWidget(self.refresh_cam_btn, 3, 0, 1, 2)

        self.start_cam_btn = QPushButton("Start Cam")
        self.start_cam_btn.clicked.connect(self.start_camera)
        self.stop_cam_btn = QPushButton("Stop Cam")
        self.stop_cam_btn.clicked.connect(self.stop_camera)
        self.stop_cam_btn.setEnabled(False)
        layout.addWidget(self.start_cam_btn, 4, 0)
        layout.addWidget(self.stop_cam_btn, 4, 1)

        return group

    def _build_flies_group(self):
        group = QGroupBox("Flies")
        layout = QGridLayout(group)

        layout.addWidget(QLabel("Number of Flies"), 0, 0)
        self.num_flies_spin = QSpinBox()
        self.num_flies_spin.setRange(1, MAX_FLIES)
        self.num_flies_spin.setValue(DEFAULT_NUM_FLIES)
        layout.addWidget(self.num_flies_spin, 0, 1)

        self.apply_num_flies_btn = QPushButton("Apply")
        self.apply_num_flies_btn.clicked.connect(self.apply_num_flies)
        layout.addWidget(self.apply_num_flies_btn, 0, 2)

        return group

    def _build_roi_group(self):
        group = QGroupBox("Vial ROI / Calibration")
        layout = QGridLayout(group)

        layout.addWidget(QLabel("Selected Fly"), 0, 0)
        self.selected_fly_label = QLabel("-")
        self.selected_fly_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.selected_fly_label, 0, 1)

        self.set_roi_btn = QPushButton("Set ROI")
        self.set_roi_btn.clicked.connect(self.arm_roi_selection)
        layout.addWidget(self.set_roi_btn, 1, 0)

        self.clear_roi_btn = QPushButton("Clear ROI")
        self.clear_roi_btn.clicked.connect(self.clear_roi)
        layout.addWidget(self.clear_roi_btn, 1, 1)

        layout.addWidget(QLabel("Vial Length (mm)"), 2, 0)
        self.vial_w_spin = QDoubleSpinBox()
        self.vial_w_spin.setRange(1, 1000)
        self.vial_w_spin.setValue(DEFAULT_VIAL_W_MM)
        self.vial_w_spin.valueChanged.connect(self.on_vial_size_changed)
        layout.addWidget(self.vial_w_spin, 2, 1)

        layout.addWidget(QLabel("Vial Width (mm)"), 3, 0)
        self.vial_h_spin = QDoubleSpinBox()
        self.vial_h_spin.setRange(1, 1000)
        self.vial_h_spin.setValue(DEFAULT_VIAL_H_MM)
        self.vial_h_spin.valueChanged.connect(self.on_vial_size_changed)
        layout.addWidget(self.vial_h_spin, 3, 1)

        layout.addWidget(QLabel("Grid Divisions"), 4, 0)
        self.grid_div_spin = QSpinBox()
        self.grid_div_spin.setRange(0, 20)
        self.grid_div_spin.setValue(DEFAULT_GRID_DIVISIONS)
        self.grid_div_spin.setToolTip("Secondary reference grid drawn inside each ROI (0 = off)")
        self.grid_div_spin.valueChanged.connect(self.engine.set_grid_divisions)
        layout.addWidget(self.grid_div_spin, 4, 1)
        layout.addWidget(QLabel("Fly Color"), 6, 0)
        color_row = QHBoxLayout()
        self.fly_color_btn = QPushButton("Pick Color")
        self.fly_color_btn.clicked.connect(self.on_fly_color_clicked)
        self.fly_color_swatch = QLabel()
        self.fly_color_swatch.setFixedSize(24, 24)
        self.fly_color_swatch.setStyleSheet("background-color: #000000; border: 1px solid #888;")
        color_row.addWidget(self.fly_color_btn)
        color_row.addWidget(self.fly_color_swatch)
        layout.addLayout(color_row, 6, 1)

        self.save_layout_btn = QPushButton("Save Layout")
        self.save_layout_btn.clicked.connect(self.save_layout)
        layout.addWidget(self.save_layout_btn, 5, 0)

        self.load_layout_btn = QPushButton("Load Layout")
        self.load_layout_btn.clicked.connect(self.load_layout)
        layout.addWidget(self.load_layout_btn, 5, 1)

        return group

    def _build_detection_group(self):
        group = QGroupBox("Detection")
        layout = QGridLayout(group)

        layout.addWidget(QLabel("Threshold"), 0, 0)
        self.thresh_slider = QSlider(Qt.Horizontal)
        self.thresh_slider.setRange(0, 255)
        self.thresh_slider.setValue(self.engine.threshold)
        self.thresh_slider.valueChanged.connect(self.on_threshold_changed)
        self.thresh_value_label = QLabel(str(self.engine.threshold))
        layout.addWidget(self.thresh_slider, 0, 1)
        layout.addWidget(self.thresh_value_label, 0, 2)

        layout.addWidget(QLabel("Min Size"), 1, 0)
        self.min_size_spin = QSpinBox()
        self.min_size_spin.setRange(1, 500)
        self.min_size_spin.setValue(self.engine.min_blob_area)
        self.min_size_spin.valueChanged.connect(self.engine.set_min_blob_area)
        layout.addWidget(self.min_size_spin, 1, 1)

        layout.addWidget(QLabel("Smoothing"), 2, 0)
        self.smoothing_spin = QSpinBox()
        self.smoothing_spin.setRange(1, 20)
        self.smoothing_spin.setValue(self.engine.smoothing_window)
        self.smoothing_spin.valueChanged.connect(self.engine.set_smoothing_window)
        layout.addWidget(self.smoothing_spin, 2, 1)

        self.preview_mask_check = QCheckBox("Preview Mask")
        self.preview_mask_check.stateChanged.connect(
            lambda s: self.engine.set_preview_mask(s == Qt.Checked))
        layout.addWidget(self.preview_mask_check, 3, 0, 1, 2)

        return group

    def _build_session_group(self):
        group = QGroupBox("Session")
        layout = QGridLayout(group)

        layout.addWidget(QLabel("Session Name"), 0, 0)
        self.session_name_edit = QLineEdit("session1")
        layout.addWidget(self.session_name_edit, 0, 1)

        layout.addWidget(QLabel("Log Interval (s)"), 1, 0)
        self.log_interval_spin = QDoubleSpinBox()
        self.log_interval_spin.setRange(0.1, 60.0)
        self.log_interval_spin.setSingleStep(0.5)
        self.log_interval_spin.setValue(DEFAULT_LOG_INTERVAL_SEC)
        self.log_interval_spin.valueChanged.connect(self.engine.set_log_interval)
        layout.addWidget(self.log_interval_spin, 1, 1)

        layout.addWidget(QLabel("Duration (min, 0=unlimited)"), 2, 0)
        self.duration_spin = QSpinBox()
        self.duration_spin.setRange(0, 1440)
        self.duration_spin.setValue(0)
        self.duration_spin.valueChanged.connect(
            lambda v: setattr(self, "duration_limit_sec", v * 60))
        layout.addWidget(self.duration_spin, 2, 1)

        self.start_track_btn = QPushButton("Start Tracking")
        self.start_track_btn.clicked.connect(self.start_tracking)
        layout.addWidget(self.start_track_btn, 3, 0)

        self.stop_track_btn = QPushButton("Stop Tracking")
        self.stop_track_btn.clicked.connect(self.stop_tracking)
        self.stop_track_btn.setEnabled(False)
        layout.addWidget(self.stop_track_btn, 3, 1)

        self.reset_data_btn = QPushButton("Reset Data")
        self.reset_data_btn.clicked.connect(self.reset_data)
        layout.addWidget(self.reset_data_btn, 4, 0, 1, 2)

        return group

    def _build_export_group(self):
        group = QGroupBox("Export")
        layout = QGridLayout(group)

        self.export_folder_btn = QPushButton("Export Folder")
        self.export_folder_btn.clicked.connect(self.choose_export_folder)
        layout.addWidget(self.export_folder_btn, 0, 0)
        self.export_folder_label = QLabel(self.export_folder)
        self.export_folder_label.setWordWrap(True)
        layout.addWidget(self.export_folder_label, 0, 1)

        self.export_excel_btn = QPushButton("Save Excel")
        self.export_excel_btn.clicked.connect(self.export_excel)
        layout.addWidget(self.export_excel_btn, 1, 0)

        self.export_traj_btn = QPushButton("Export Paths")
        self.export_traj_btn.clicked.connect(self.export_trajectories)
        layout.addWidget(self.export_traj_btn, 1, 1)

        layout.addWidget(QLabel("Movement Threshold (mm)"), 2, 0)
        thresh_row = QHBoxLayout()
        self.movement_thresh_slider = QSlider(Qt.Horizontal)
        self.movement_thresh_slider.setRange(0, 50)  # represents 0.0-5.0 mm in 0.1 steps
        self.movement_thresh_slider.setValue(int(DEFAULT_MOVEMENT_THRESHOLD_MM * 10))
        self.movement_thresh_slider.valueChanged.connect(self.on_movement_threshold_changed)
        self.movement_thresh_label = QLabel(f"{DEFAULT_MOVEMENT_THRESHOLD_MM:.1f} mm")
        thresh_row.addWidget(self.movement_thresh_slider)
        thresh_row.addWidget(self.movement_thresh_label)
        layout.addLayout(thresh_row, 2, 1)

        self.auto_save_check = QCheckBox("Auto-Save")
        self.auto_save_check.stateChanged.connect(self.on_auto_save_toggled)
        layout.addWidget(self.auto_save_check, 3, 0)

        layout.addWidget(QLabel("Auto-Save Every (min)"), 4, 0)
        self.auto_save_interval_spin = QSpinBox()
        self.auto_save_interval_spin.setRange(1, 60)
        self.auto_save_interval_spin.setValue(5)
        layout.addWidget(self.auto_save_interval_spin, 4, 1)

        return group

    # -- Camera discovery ---------------------------------------------------
    def refresh_camera_list(self):
        was_live = self.timer.isActive()
        if was_live:
            self.stop_camera()

        self.camera_select.clear()
        found = TrackerEngine.detect_cameras()  # list of (index, name)
        if found:
            for idx, name in found:
                self.camera_select.addItem(name, userData=idx)
            self.start_cam_btn.setEnabled(True)
        else:
            self.camera_select.addItem("No camera detected", userData=None)
            self.start_cam_btn.setEnabled(False)

    # -- Fly count / naming ----------------------------------------------------
    def apply_num_flies(self):
        n = self.num_flies_spin.value()
        self.engine.set_num_flies(n)
        self._selected_fly_idx = 0
        self.rebuild_status_table()

    def rebuild_status_table(self):
        self._suppress_table_signal = True
        n = self.engine.num_flies
        self.status_table.setRowCount(n)
        for i, roi in enumerate(self.engine.rois):
            name_item = QTableWidgetItem(roi.name)
            name_item.setFlags(name_item.flags() | Qt.ItemIsEditable)
            self.status_table.setItem(i, 0, name_item)

            det_item = QTableWidgetItem("-")
            det_item.setFlags(det_item.flags() & ~Qt.ItemIsEditable)
            self.status_table.setItem(i, 1, det_item)

            pos_item = QTableWidgetItem("-")
            pos_item.setFlags(pos_item.flags() & ~Qt.ItemIsEditable)
            self.status_table.setItem(i, 2, pos_item)

            dist_item = QTableWidgetItem("0.00")
            dist_item.setFlags(dist_item.flags() & ~Qt.ItemIsEditable)
            self.status_table.setItem(i, 3, dist_item)
        self._suppress_table_signal = False

        if n:
            self._selected_fly_idx = min(self._selected_fly_idx, n - 1)
            self.status_table.selectRow(self._selected_fly_idx)
            self.on_fly_selected(self._selected_fly_idx)

    def on_table_item_changed(self, item):
        if self._suppress_table_signal or item.column() != 0:
            return
        idx = item.row()
        if idx >= self.engine.num_flies:
            return
        self.engine.set_fly_name(idx, item.text())
        # reflect any normalization (e.g. blank -> default name) back into the cell
        self._suppress_table_signal = True
        item.setText(self.engine.rois[idx].name)
        self._suppress_table_signal = False
        if idx == self._selected_fly_idx:
            self.selected_fly_label.setText(self.engine.rois[idx].name)

    def on_table_selection_changed(self):
        rows = self.status_table.selectionModel().selectedRows()
        if not rows:
            return
        idx = rows[0].row()
        self.on_fly_selected(idx)

    def on_fly_selected(self, idx):
        if idx < 0 or idx >= self.engine.num_flies:
            return
        self._selected_fly_idx = idx
        roi = self.engine.rois[idx]
        self.selected_fly_label.setText(roi.name)
        self.vial_w_spin.blockSignals(True)
        self.vial_h_spin.blockSignals(True)
        self.vial_w_spin.setValue(roi.real_w_mm)
        self.vial_h_spin.setValue(roi.real_h_mm)
        self.vial_w_spin.blockSignals(False)
        self.vial_h_spin.blockSignals(False)
        current_color = roi.color if roi.color else fly_color(idx)
        self.fly_color_swatch.setStyleSheet(f"background-color: {current_color}; border: 1px solid #888;")

    def on_fly_color_clicked(self):
        idx = self._selected_fly_idx
        roi = self.engine.rois[idx]
        current = QColor(roi.color if roi.color else fly_color(idx))
        picked = QColorDialog.getColor(current, self, "Pick Fly Color")
        if picked.isValid():
            hex_color = picked.name()
            self.engine.set_fly_color(idx, hex_color)
            self.fly_color_swatch.setStyleSheet(f"background-color: {hex_color}; border: 1px solid #888;")

    # -- Mode switching -------------------------------------------------------
    def on_mode_changed(self):
        live = self.live_mode_radio.isChecked()
        has_camera = self.camera_select.count() > 0 and self.camera_select.currentData() is not None
        self.start_cam_btn.setEnabled(live and has_camera)
        self.stop_cam_btn.setEnabled(False)
        self.camera_select.setEnabled(live)
        self.refresh_cam_btn.setEnabled(live)
        self.load_video_btn.setEnabled(not live)
        self.process_video_btn.setEnabled(not live and self.loaded_video_path is not None)
        self.start_track_btn.setEnabled(live)
        self.stop_track_btn.setEnabled(False)

    # -- Camera lifecycle -----------------------------------------------------
    def start_camera(self):
        cam_idx = self.camera_select.currentData()
        if cam_idx is None:
            QMessageBox.warning(self, "No Camera", "No camera detected. Click 'Refresh Cameras' after connecting one.")
            return
        _, want_w, want_h = RESOLUTIONS[self.resolution_select.currentIndex()]
        want_fps = FPS_OPTIONS[self.fps_select.currentIndex()]
        try:
            got_w, got_h, got_fps = self.engine.open_video_source(cam_idx, width=want_w, height=want_h, fps=want_fps)
        except RuntimeError as e:
            QMessageBox.critical(self, "Camera Error", str(e))
            return
        if (got_w, got_h) != (want_w, want_h):
            QMessageBox.information(
                self, "Resolution",
                f"Requested {want_w}x{want_h}, but the camera/driver granted {got_w}x{got_h} instead.\n"
                "This is a camera/driver limitation, not an app setting."
            )
        # Poll at the camera's own requested frame rate rather than a fixed
        # 100ms (10fps) tick, so the preview isn't artificially capped below
        # what the camera can actually deliver.
        interval_ms = max(1, int(1000 / (got_fps if got_fps and got_fps > 0 else want_fps)))
        self.timer.start(interval_ms)
        self.start_cam_btn.setEnabled(False)
        self.stop_cam_btn.setEnabled(True)

    def stop_camera(self):
        self.timer.stop()
        self.engine.close_video_source()
        has_camera = self.camera_select.currentData() is not None
        self.start_cam_btn.setEnabled(has_camera)
        self.stop_cam_btn.setEnabled(False)

    # -- Video file mode --------------------------------------------------------
    def load_video_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Load Video", "", "Video Files (*.mp4 *.avi *.mov *.mkv)")
        if path:
            self.loaded_video_path = path
            self.video_path_label.setText(os.path.basename(path))
            self.process_video_btn.setEnabled(True)

    def process_video(self):
        if not self.loaded_video_path:
            return
        self.process_video_btn.setEnabled(False)
        self.load_video_btn.setEnabled(False)
        try:
            self._batch_generator = self.engine.process_video_batch(self.loaded_video_path)
            for _ in self._batch_generator:
                QApplication.processEvents()  # keep the GUI responsive during the batch loop
        except RuntimeError as e:
            QMessageBox.critical(self, "Video Error", str(e))
        finally:
            self.process_video_btn.setEnabled(True)
            self.load_video_btn.setEnabled(True)
            QMessageBox.information(self, "Done", "Video processing complete. You can now export results.")

    # -- ROI setup --------------------------------------------------------
    def arm_roi_selection(self):
        self.video_label.arm_roi_mode(True)

    def on_roi_drawn(self, rect):
        self.engine.set_roi(self._selected_fly_idx, rect)

    def clear_roi(self):
        self.engine.clear_roi(self._selected_fly_idx)

    def on_vial_size_changed(self):
        self.engine.set_vial_size(self._selected_fly_idx, self.vial_w_spin.value(), self.vial_h_spin.value())

    def save_layout(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Layout", "layout.json", "JSON Files (*.json)")
        if path:
            self.engine.save_layout(path)
            QMessageBox.information(self, "Saved", f"Layout saved to {path}")

    def load_layout(self):
        path, _ = QFileDialog.getOpenFileName(self, "Load Layout", "", "JSON Files (*.json)")
        if path:
            self.engine.load_layout(path)
            self.num_flies_spin.setValue(self.engine.num_flies)
            self.thresh_slider.setValue(self.engine.threshold)
            self.min_size_spin.setValue(self.engine.min_blob_area)
            self.log_interval_spin.setValue(self.engine.log_interval_sec)
            self.smoothing_spin.setValue(self.engine.smoothing_window)
            self._selected_fly_idx = 0
            self.rebuild_status_table()
            QMessageBox.information(self, "Loaded", f"Layout loaded from {path}")

    # -- Detection ----------------------------------------------------------
    def on_threshold_changed(self, value):
        self.engine.set_threshold(value)
        self.thresh_value_label.setText(str(value))

    def on_movement_threshold_changed(self, slider_value):
        mm = slider_value / 10.0
        self.engine.set_movement_threshold(mm)
        self.movement_thresh_label.setText(f"{mm:.1f} mm")

    # -- Tracking / session (live mode) -------------------------------------------
    def start_tracking(self):
        self.engine.tracking_active = True
        self.engine.last_log_time = 0.0
        self.session_start_time = time.time()
        self.elapsed_timer.start(1000)
        self.start_track_btn.setEnabled(False)
        self.stop_track_btn.setEnabled(True)

    def stop_tracking(self):
        self.engine.tracking_active = False
        self.elapsed_timer.stop()
        self.start_track_btn.setEnabled(True)
        self.stop_track_btn.setEnabled(False)

    def reset_data(self):
        self.engine.reset_data()
        self.session_start_time = None
        self.elapsed_label.setText("Elapsed Time: 00:00:00")
        self.progress_bar.setValue(0)
        self.rebuild_status_table()

    def update_elapsed(self):
        if self.session_start_time is None:
            return
        elapsed = int(time.time() - self.session_start_time)
        h, rem = divmod(elapsed, 3600)
        m, s = divmod(rem, 60)
        self.elapsed_label.setText(f"Elapsed Time: {h:02d}:{m:02d}:{s:02d}")

        if self.duration_limit_sec and elapsed >= self.duration_limit_sec:
            self.stop_tracking()
            QMessageBox.information(self, "Session Complete", "Duration limit reached — tracking stopped.")

        if self.auto_save_enabled:
            interval_sec = self.auto_save_interval_spin.value() * 60
            if elapsed and elapsed % interval_sec == 0:
                self._do_auto_save()

    def on_auto_save_toggled(self, state):
        self.auto_save_enabled = (state == Qt.Checked)

    def _do_auto_save(self):
        name = self.session_name_edit.text().strip() or "session"
        path = os.path.join(self.export_folder, f"{name}_autosave.xlsx")
        self.engine.export_excel(path)

    # -- Frame / status / progress display -----------------------------------------
    def on_frame(self, frame):
        self.video_label.frame_size = (frame.shape[1], frame.shape[0])
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb.shape
        qimg = QImage(rgb.data, w, h, ch * w, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(qimg).scaled(
            self.video_label.width(), self.video_label.height(), Qt.KeepAspectRatio
        )
        self.video_label.setPixmap(pixmap)

    def on_status(self, fly_idx, pos_mm, detected):
        if fly_idx >= self.status_table.rowCount():
            return
        self._suppress_table_signal = True
        self.status_table.item(fly_idx, 1).setText("Yes" if detected else "No")
        text = f"{pos_mm[0]:.1f}, {pos_mm[1]:.1f}" if pos_mm else "-"
        self.status_table.item(fly_idx, 2).setText(text)
        self._suppress_table_signal = False

    def on_distance_updated(self, fly_idx, total_distance_mm):
        if fly_idx >= self.status_table.rowCount():
            return
        self._suppress_table_signal = True
        self.status_table.item(fly_idx, 3).setText(f"{total_distance_mm:.2f}")
        self._suppress_table_signal = False

    def on_fps(self, fps):
        self.fps_label.setText(f"Frame Rate: {fps:.1f}")

    def on_progress(self, pct):
        self.progress_bar.setValue(pct)

    # -- Export ---------------------------------------------------------------
    def choose_export_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Export Folder", self.export_folder)
        if folder:
            self.export_folder = folder
            self.export_folder_label.setText(folder)

    def _session_timestamp_prefix(self):
        session_name = self.session_name_edit.text().strip() or "session"
        now = datetime.datetime.now()
        return f"{session_name}_{now.strftime('%Y%m%d')}_{now.strftime('%H%M%S')}"

    def export_excel(self):
        prefix = self._session_timestamp_prefix()
        path = os.path.join(self.export_folder, f"{prefix}.xlsx")
        self.engine.export_excel(path)
        QMessageBox.information(self, "Saved", f"Data saved to {path}")

    def export_trajectories(self):
        prefix = self._session_timestamp_prefix()
        self.engine.export_trajectories(self.export_folder, prefix)
        QMessageBox.information(self, "Saved", f"Trajectory images saved to {self.export_folder}")

    def closeEvent(self, event):
        self.stop_camera()
        event.accept()


def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.resize(1200, 720)
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
